"""
Lectura y escritura del Excel maestro.
Extrae productos existentes y permite actualizarlos.
"""
import openpyxl
from copy import copy
import sys

sys.stdout.reconfigure(encoding='utf-8')


class ExcelMaestro:
    """Maneja la lectura y escritura del TechnoStore.xlsx."""
    
    def __init__(self, ruta_excel, config_path='config.json'):
        import json
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        self.ruta_excel = ruta_excel
        self.wb = None
    
    def __enter__(self):
        self.wb = openpyxl.load_workbook(self.ruta_excel)
        return self
    
    def __exit__(self, *args):
        if self.wb:
            self.wb.close()
    
    def leer_pestana(self, nombre_pestana):
        """
        Lee una pestaña del Excel y devuelve los productos.
        
        Returns:
            list[dict]: productos con sus datos y posición (fila)
        """
        if nombre_pestana not in self.wb.sheetnames:
            return []
        
        ws = self.wb[nombre_pestana]
        
        # Buscar fila de encabezado
        header_row = self._find_header_row(ws)
        if not header_row:
            return []
        
        # Buscar columnas
        cols = self._find_columns(ws, header_row)
        
        productos = []
        for row_num in range(header_row + 1, ws.max_row + 1):
            row_data = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
            
            # Verificar si la fila está vacía o es un subtítulo
            valores = [str(v).strip() if v is not None else '' for v in row_data]
            if not any(valores):
                continue
            
            # Saltar filas que son subtítulos (ej: "MÓDULOS SAMSUNG")
            primer_valor = valores[0] if valores else ''
            if self._es_subtitulo(primer_valor):
                continue
            
            producto = self._extraer_producto(row_data, cols, row_num)
            if producto:
                productos.append(producto)
        
        return productos
    
    def _find_header_row(self, ws):
        """Busca la fila de encabezados."""
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val:
                    val_str = str(val).upper().strip()
                    if 'MARCA' in val_str or 'MODELO' in val_str:
                        return row
        return None
    
    def _find_columns(self, ws, header_row):
        """Mapea nombres de columnas a números de columna."""
        cols = {}
        
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col).value).strip().upper() if ws.cell(row=header_row, column=col).value else ''
            
            if 'MARCA' in val:
                cols['marca'] = col
            elif 'MODELO' in val or 'DESCRIPCION' in val:
                cols['modelo'] = col
            elif 'CALIDAD' in val or 'TIPO' in val or 'COLOR' in val:
                cols['calidad'] = col
            elif 'PRECIO PESOS' in val or 'PESOS' in val:
                cols['precio_pesos'] = col
            elif 'PRECIO DOLAR' in val or 'DOLAR' in val or 'U$S' in val:
                cols['precio_dolar'] = col
            elif 'VENTA' in val:
                cols['venta'] = col
            elif 'STOCK' in val:
                cols['stock'] = col
        
        return cols
    
    def _es_subtitulo(self, texto):
        """Detecta si una fila es un subtítulo/categoría, no un producto."""
        marcadores = ['MÓDULO', 'BATERIA', 'COMPONENTE', 'TAPA', 'PLACA',
                      'ACCESORIO', 'REPUESTO', 'MODELO', 'LISTA']
        
        texto_upper = texto.upper().strip()
        
        # Si empieza con una de estas palabras y no tiene otros datos
        for m in marcadores:
            if texto_upper.startswith(m) and len(texto_upper) < 40:
                # Verificar si parece título (no es marca+modelo)
                palabras = texto_upper.split()
                if len(palabras) <= 3:
                    return True
        
        return False
    
    def _extraer_producto(self, row_data, cols, row_num):
        """Extrae datos de una fila del Excel."""
        def get_val(col_num):
            if col_num and col_num <= len(row_data):
                v = row_data[col_num - 1]
                if v is None:
                    return ''
                return str(v).strip()
            return ''
        
        def get_num(col_num):
            v = get_val(col_num)
            if not v:
                return None
            try:
                return float(v.replace(',', '.').replace('$', '').replace(' ', ''))
            except:
                return None
        
        marca = get_val(cols.get('marca', 1))
        modelo = get_val(cols.get('modelo', 2))
        calidad = get_val(cols.get('calidad', 3))
        precio_pesos = get_num(cols.get('precio_pesos'))
        precio_dolar = get_num(cols.get('precio_dolar'))
        venta = get_num(cols.get('venta'))
        stock = get_num(cols.get('stock'))

        # Si el modelo tiene todo junto (ej: "SAMSUNG J1 ACE OLED2"),
        # extraer marca y calidad para que matchee con el PDF
        if modelo and not calidad:
            marcas = ['SAMSUNG', 'MOTOROLA', 'IPHONE', 'HUAWEI', 'LG', 'SONY',
                     'NOKIA', 'TCL', 'ALCATEL', 'XIAOMI', 'ZTE', 'HONOR',
                     'INFINIX', 'REALME', 'OPPO', 'ONEPLUS', 'TECNO', 'NUBIA',
                     'APPLE', 'BLADE']
            calidades = ['ORIGINAL', 'OLED', 'OLED2', 'INCELL', 'AMOLED',
                         'SUPER AMOLED', 'C/MARCO', 'S/M', 'S/MARCO',
                         'CON MARCO', 'SIN MARCO', 'PREMIUN', 'PREMIUM',
                         'SVC', 'AMERICANO']

            modelo_upper = modelo.upper()
            partes = modelo_upper.split()

            # Extraer marca del inicio si está
            for m in marcas:
                if modelo_upper.startswith(m + ' ') or modelo_upper == m:
                    if not marca:
                        marca = m
                    # Sacar la marca del modelo
                    idx = modelo_upper.find(m)
                    modelo = modelo[idx + len(m):].strip()
                    modelo_upper = modelo.upper()
                    break

            # Extraer calidad del final
            for c in sorted(calidades, key=len, reverse=True):
                if modelo_upper.endswith(' ' + c) or modelo_upper.endswith(c):
                    calidad = c
                    modelo = modelo[:-(len(c))].strip()
                    break

        # Al menos debe tener modelo y algún precio
        if not modelo and not marca:
            return None
        
        if precio_pesos is None and precio_dolar is None:
            # Quizás el precio está en la columna de "calidad" (como vimos en PGVJ REPUESTOS)
            calidad_precio = get_num(cols.get('calidad'))
            if calidad_precio and calidad_precio > 0:
                precio_pesos = calidad_precio
                calidad = ''
            else:
                return None
        
        return {
            'fila': row_num,
            'marca': marca,
            'modelo': modelo,
            'calidad_o_color': calidad,
            'precio_pesos': precio_pesos,
            'precio_dolar': precio_dolar,
            'precio_venta': venta,
            'stock': int(stock) if stock else None
        }
    
    def actualizar_producto(self, nombre_pestana, fila, columna, valor):
        """Actualiza una celda específica."""
        if nombre_pestana not in self.wb.sheetnames:
            return False
        
        ws = self.wb[nombre_pestana]
        ws.cell(row=fila, column=columna).value = valor
        return True
    
    def agregar_producto(self, nombre_pestana, producto):
        """
        Agrega un nuevo producto al final de una pestaña.
        
        Args:
            nombre_pestana: nombre de la pestaña
            producto: dict con marca, modelo, calidad_o_color, precio, moneda, stock
        """
        if nombre_pestana not in self.wb.sheetnames:
            return False
        
        ws = self.wb[nombre_pestana]
        header_row = self._find_header_row(ws)
        cols = self._find_columns(ws, header_row)
        
        # Buscar primera fila vacía
        nueva_fila = ws.max_row + 1
        
        # Escribir datos
        if 'marca' in cols:
            ws.cell(row=nueva_fila, column=cols['marca']).value = producto.get('marca', '')
        if 'modelo' in cols:
            ws.cell(row=nueva_fila, column=cols['modelo']).value = producto.get('modelo', '')
        if 'calidad' in cols:
            ws.cell(row=nueva_fila, column=cols['calidad']).value = producto.get('calidad_o_color', '')
        
        precio = producto.get('precio')
        moneda = producto.get('moneda', 'ARS')
        
        if moneda == 'USD' and 'precio_dolar' in cols:
            ws.cell(row=nueva_fila, column=cols['precio_dolar']).value = precio
        elif 'precio_pesos' in cols:
            ws.cell(row=nueva_fila, column=cols['precio_pesos']).value = precio
        
        stock = producto.get('stock')
        if stock and 'stock' in cols:
            ws.cell(row=nueva_fila, column=cols['stock']).value = stock
        
        return True
    
    def guardar(self, ruta_salida=None):
        """Guarda los cambios en el archivo."""
        ruta = ruta_salida or self.ruta_excel
        self.wb.save(ruta)
        return True
    
    def obtener_pestanas_proveedor(self, proveedor):
        """Devuelve las pestañas asociadas a un proveedor."""
        return self.config.get('proveedores', {}).get(proveedor, {}).get('pestanas', [])
