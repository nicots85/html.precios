"""
Interfaz IA proveedor-agnóstico.
Soporta cualquier API compatible con OpenAI (OpenAI, Anthropic vía proxy, etc.)
"""
import json
import os
import base64
import sys

sys.stdout.reconfigure(encoding='utf-8')


class AIProvider:
    """Interfaz unificada para distintos proveedores de IA."""

    def __init__(self, config_path='config.json'):
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        ai_config = config.get('ai', {})
        self.provider = ai_config.get('provider', 'openai')
        self.api_key = ai_config.get('api_key', '')
        self.api_base = ai_config.get('api_base', 'https://api.openai.com/v1')
        self.model = ai_config.get('model', 'gpt-4o')
        self.max_tokens = ai_config.get('max_tokens', 4000)
        self.temperature = ai_config.get('temperature', 0)
        self.mock_mode = ai_config.get('mock_mode', False)  # MODO TEST

    def interpretar_archivo(self, ruta_archivo, tipo_archivo):
        """
        Interpreta un archivo (xlsx, pdf, imagen, texto) y devuelve
        una lista de productos en estructura común.
        
        Args:
            ruta_archivo: path al archivo
            tipo_archivo: 'excel', 'pdf', 'imagen', 'texto'
        
        Returns:
            list[dict]: lista de productos normalizados
        """
        if self.mock_mode:
            print("🧪 MODO MOCK: Simulando IA sin llamada real")
            return self._mock_interpretar(ruta_archivo, tipo_archivo)
        
        if tipo_archivo == 'excel':
            return self._interpretar_excel(ruta_archivo)
        elif tipo_archivo == 'pdf':
            return self._interpretar_pdf(ruta_archivo)
        elif tipo_archivo == 'imagen':
            return self._interpretar_imagen(ruta_archivo)
        elif tipo_archivo == 'texto':
            return self._interpretar_texto(ruta_archivo)
        else:
            raise ValueError(f"Tipo no soportado: {tipo_archivo}")

    def _interpretar_excel(self, ruta_archivo):
        """Interpreta un Excel directamente con openpyxl + IA si hace falta."""
        import openpyxl
        
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        productos = []
        
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            filas = []
            
            for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=True):
                fila = [str(c) if c is not None else '' for c in row]
                if any(fila):
                    filas.append(fila)
            
            if filas:
                prompt = self._armar_prompt_excel(filas, hoja)
                resultado = self._consultar_ia(prompt)
                if resultado:
                    productos.extend(resultado)
        
        return productos

    def _interpretar_pdf(self, ruta_archivo):
        """Interpreta un PDF usando la IA directamente."""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(ruta_archivo)
            texto_completo = ""
            
            for pagina in doc:
                texto_completo += pagina.get_text() + "\n---\n"
            
            if texto_completo.strip():
                prompt = self._armar_prompt_texto(texto_completo, "PDF")
                return self._consultar_ia(prompt) or []
            
            # Si no se pudo extraer texto, mandar como imagen
            return self._interpretar_pdf_como_imagen(doc)
            
        except ImportError:
            return self._interpretar_pdf_como_imagen(ruta_archivo)

    def _interpretar_pdf_como_imagen(self, doc_o_ruta):
        """Fallback: convierte páginas a imagen y usa visión."""
        try:
            import fitz
            if isinstance(doc_o_ruta, str):
                doc = fitz.open(doc_o_ruta)
            else:
                doc = doc_o_ruta
            
            pagina = doc[0]
            pix = pix = pagina.get_pixmap(dpi=200)
            img_bytes = pix.tobytes("png")
            
            return self._consultar_ia_vision(img_bytes, "Extraé los productos de esta imagen de una lista de precios.")
        except Exception as e:
            print(f"Error interpretando PDF como imagen: {e}")
            return []

    def _interpretar_imagen(self, ruta_archivo):
        """Interpreta una imagen (foto de lista impresa)."""
        with open(ruta_archivo, 'rb') as f:
            img_bytes = f.read()
        
        return self._consultar_ia_vision(img_bytes, "Extraé los productos de esta imagen de una lista de precios.") or []

    def _interpretar_texto(self, ruta_archivo):
        """Interpreta texto suelto."""
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            texto = f.read()
        
        prompt = self._armar_prompt_texto(texto, "texto")
        return self._consultar_ia(prompt) or []

    def _armar_prompt_excel(self, filas, nombre_hoja):
        """Arma el prompt para interpretar datos de Excel."""
        texto = f"HOJA: {nombre_hoja}\n\n"
        texto += "Filas (separadas por |):\n"
        for i, fila in enumerate(filas[:30]):
            texto += f"F{i}: {' | '.join(fila)}\n"
        
        return self._prompt_base(texto)

    def _armar_prompt_texto(self, texto, origen):
        """Arma el prompt para interpretar texto/PDF."""
        return self._prompt_base(f"ORIGEN: {origen}\n\n{texto[:5000]}")

    def _prompt_base(self, contenido):
        """Prompt base del sistema."""
        return f"""Sos un asistente que extrae productos de listas de precios de proveedores de electrónica.

Devolvé SOLO un array JSON con los productos encontrados. Cada producto debe tener:
- marca: string (ej: "SAMSUNG", "IPHONE", "MOTOROLA")
- modelo: string (ej: "A13", "EDGE 50", "15 PRO")
- calidad_o_color: string (ej: "ORIGINAL", "CON LENTE NEGRO", "OLED")
- precio: número (el precio de lista, sin $ ni comas)
- moneda: "ARS" o "USD" (deducilo del contexto)
- stock: número o null (solo si aparece explícitamente)

REGLAS:
1. Si hay columnas claras (Marca | Modelo | Calidad | Precio), usalas directamente.
2. Si el precio tiene $, asumí ARS. Si dice "U$S" o "USD", asumí USD.
3. Si no hay marca clara, deducilo del modelo (ej: "iPhone 15" → marca "IPHONE").
4. Si una fila es un subtítulo o categoría (ej: "MÓDULOS SAMSUNG"), ignorala.
5. Si una fila tiene precio 0 o está vacía, ignorala.
6. Calidad/color: si hay info de color, incluila. Si dice "ORIGINAL", "COPIA", "PREMIUM", incluilo.

EJEMPLO DE SALIDA:
[
  {{"marca": "SAMSUNG", "modelo": "A13", "calidad_o_color": "CON LENTE NEGRO", "precio": 11472, "moneda": "ARS", "stock": null}},
  {{"marca": "IPHONE", "modelo": "15 PRO", "calidad_o_color": "OLED ORIGINAL", "precio": 450, "moneda": "USD", "stock": 5}}
]

DATOS:
{contenido}

JSON:"""

    def _consultar_ia(self, prompt):
        """Consulta a la API de IA (compatible OpenAI)."""
        try:
            import urllib.request
            import urllib.error
            
            url = f"{self.api_base}/chat/completions"
            
            data = {
                "model": self.model,
                "messages": [
                    {"role": "system", "content": "Sos un asistente que extrae datos estructurados de listas de precios. Devolvé SOLO JSON válido, sin texto adicional."},
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": self.max_tokens,
                "temperature": self.temperature
            }
            
            req = urllib.request.Request(
                url,
                data=json.dumps(data).encode('utf-8'),
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {self.api_key}"
                }
            )
            
            with urllib.request.urlopen(req, timeout=120) as response:
                resultado = json.loads(response.read().decode('utf-8'))
            
            contenido = resultado['choices'][0]['message']['content'].strip()
            
            # Limpiar respuesta (a veces viene con ```json)
            if contenido.startswith('```'):
                contenido = contenido.split('\n')[1:-1]
                contenido = '\n'.join(contenido)
            
            # Parsear JSON
            try:
                productos = json.loads(contenido)
                if isinstance(productos, dict) and 'productos' in productos:
                    productos = productos['productos']
                return productos
            except json.JSONDecodeError:
                print(f"Error parseando JSON: {contenido[:200]}")
                return None
                
        except Exception as e:
            print(f"Error consultando IA: {e}")
            return None

    def _consultar_ia_vision(self, img_bytes, prompt):
        """Consulta a la IA con imagen (visión)."""
        try:
            import urllib.request
            
            url = f"{self.api_base}/chat/completions"
            
            img_b64 = base64.b64encode(img_bytes).decode('utf-8')
            
            data = {
                "model": self.model,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": self._prompt_base(prompt)},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{img_b64}"
                                }
                            }
                        ]
                    }
                ],
                "max_tokens": self.max_tokens,
                "temperature": self.temperature
            }
            
            req = urllib.request.Request(
                url,
                data=json.dumps(data).encode('utf-8'),
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {self.api_key}"
                }
            )
            
            with urllib.request.urlopen(req, timeout=180) as response:
                resultado = json.loads(response.read().decode('utf-8'))
            
            contenido = resultado['choices'][0]['message']['content'].strip()
            
            if contenido.startswith('```'):
                contenido = contenido.split('\n')[1:-1]
                contenido = '\n'.join(contenido)
            
            try:
                productos = json.loads(contenido)
                if isinstance(productos, dict) and 'productos' in productos:
                    productos = productos['productos']
                return productos
            except json.JSONDecodeError:
                print(f"Error parseando JSON de visión: {contenido[:200]}")
                return None
                
        except Exception as e:
            print(f"Error consultando IA visión: {e}")
            return None

    # ========== MODO MOCK (para testing sin IA real) ==========

    def _mock_interpretar(self, ruta_archivo, tipo_archivo):
        """Interpreta archivo en modo mock - parsea directo sin IA."""
        import openpyxl
        
        if tipo_archivo == 'excel':
            return self._mock_excel(ruta_archivo)
        elif tipo_archivo == 'pdf':
            return self._mock_pdf(ruta_archivo)
        elif tipo_archivo == 'imagen':
            return self._mock_imagen(ruta_archivo)
        elif tipo_archivo == 'texto':
            return self._mock_texto(ruta_archivo)
        else:
            raise ValueError(f"Tipo no soportado: {tipo_archivo}")

    def _mock_excel(self, ruta_archivo):
        """Parsea Excel directamente detectando columnas."""
        import openpyxl
        
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        productos = []
        
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            header_row = self._find_header_row(ws)
            if not header_row:
                continue
            
            cols = self._find_columns_mock(ws, header_row)
            if not cols.get('modelo'):
                continue
            
            for row_num in range(header_row + 1, ws.max_row + 1):
                row_data = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
                producto = self._extraer_producto_mock(row_data, cols, hoja)
                if producto:
                    productos.append(producto)
        
        return productos

    def _find_header_row(self, ws):
        for row in range(1, min(15, ws.max_row + 1)):
            for col in range(1, min(30, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val and ('MARCA' in str(val).upper() or 'MODELO' in str(val).upper() or 'PRODUCTO' in str(val).upper()):
                    return row
        return None

    def _find_columns_mock(self, ws, header_row):
        cols = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=col).value).strip().upper() if ws.cell(row=header_row, column=col).value else ''
            if 'MARCA' in val:
                cols['marca'] = col
            elif 'MODELO' in val or 'DESCRIPCION' in val:
                cols['modelo'] = col
            elif 'CALIDAD' in val or 'TIPO' in val or 'COLOR' in val:
                cols['calidad'] = col
            elif 'PRECIO PESOS' in val or 'PESOS' in val or ('PRECIO' in val and 'DOLAR' not in val):
                cols['precio_pesos'] = col
            elif 'PRECIO DOLAR' in val or 'DOLAR' in val or 'U$S' in val:
                cols['precio_dolar'] = col
            elif 'STOCK' in val:
                cols['stock'] = col
        return cols

    def _extraer_producto_mock(self, row_data, cols, hoja):
        def get_val(col_num):
            if col_num and col_num <= len(row_data):
                v = row_data[col_num - 1]
                if v is None:
                    return ''
                return str(v).strip()
            return ''
        
        def get_num(col_num):
            v = get_val(col_num)
            if not v:
                return None
            try:
                return float(v.replace(',', '.').replace('$', '').replace(' ', ''))
            except:
                return None
        
        marca = get_val(cols.get('marca'))
        modelo = get_val(cols.get('modelo'))
        calidad = get_val(cols.get('calidad'))
        precio_pesos = get_num(cols.get('precio_pesos'))
        precio_dolar = get_num(cols.get('precio_dolar'))
        stock = get_num(cols.get('stock'))
        
        if not modelo and not marca:
            return None
        
        if precio_pesos is None and precio_dolar is None:
            return None
        
        moneda = 'ARS' if precio_pesos else 'USD'
        precio = precio_pesos or precio_dolar
        
        return {
            'marca': marca or '',
            'modelo': modelo,
            'calidad_o_color': calidad or '',
            'precio': precio,
            'moneda': moneda,
            'stock': int(stock) if stock else None
        }

    def _mock_pdf(self, ruta_archivo):
        try:
            import fitz
            doc = fitz.open(ruta_archivo)
            texto = ""
            for pagina in doc:
                texto += pagina.get_text() + "\n"
            return self._mock_texto_from_string(texto)
        except:
            return []

    def _mock_imagen(self, ruta_archivo):
        return []

    def _mock_texto(self, ruta_archivo):
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            texto = f.read()
        return self._mock_texto_from_string(texto)

    def _mock_texto_from_string(self, texto):
        """Parsea texto suelto línea por línea buscando patrones precio."""
        import re
        productos = []
        lineas = texto.split('\n')
        
        for linea in lineas:
            linea = linea.strip()
            if not linea:
                continue
            
            # Buscar patrón: algo $numero o algo USD numero
            match = re.search(r'(.+?)\s+[\$\$]?\s*([\d.,]+)\s*(USD|ARS|U\$S)?', linea, re.IGNORECASE)
            if match:
                desc = match.group(1).strip()
                precio_str = match.group(2).replace(',', '.').replace('.', '', match.group(2).count('.')-1)
                try:
                    precio = float(precio_str)
                    moneda = (match.group(3) or 'ARS').upper()
                    if moneda in ['USD', 'U$S', 'U$S']:
                        moneda = 'USD'
                    else:
                        moneda = 'ARS'
                    
                    # Separar marca/modelo de descripción
                    partes = desc.split()
                    marca = partes[0] if partes else ''
                    modelo = ' '.join(partes[1:]) if len(partes) > 1 else ''
                    
                    productos.append({
                        'marca': marca.upper(),
                        'modelo': modelo.upper(),
                        'calidad_o_color': '',
                        'precio': precio,
                        'moneda': moneda,
                        'stock': None
                    })
                except:
                    pass
        return productos
