"""
Generador de productos.json desde TechnoStore.xlsx.
Lee el Excel maestro y produce el catálogo que consume la web (sin costos ni márgenes).

Uso:
    python backend/generar_productos.py

Regenera productos.json en la raíz del proyecto (lo que se sube a Firebase Hosting).
"""
import json
import os
import re
import sys
import unicodedata
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')


def leer_excel(ruta_excel):
    """Lee el Excel maestro y devuelve productos agrupados por pestaña."""
    import openpyxl

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    productos = []

    for tab_name in wb.sheetnames:
        ws = wb[tab_name]

        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val and ('MARCA' in str(val).upper() or 'MODELO' in str(val).upper()):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            continue

        cols = {}
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=c).value).strip().upper() if ws.cell(row=header_row, column=c).value else ''
            if 'MARCA' in val:
                cols['marca'] = c
            elif 'MODELO' in val or 'DESCRIPCION' in val:
                cols['modelo'] = c
            elif 'CALIDAD' in val or 'TIPO' in val:
                cols['calidad'] = c
            elif 'VENTA' in val:
                cols['venta'] = c
            elif 'STOCK' in val:
                cols['stock'] = c

        if not cols.get('venta'):
            continue

        proveedor = detectar_proveedor(tab_name)
        categoria = detectar_categoria(tab_name)

        for row_num in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]

            marca = str(row[cols['marca'] - 1] or '').strip() if cols.get('marca') else ''
            modelo = str(row[cols['modelo'] - 1] or '').strip() if cols.get('modelo') else ''
            calidad = str(row[cols['calidad'] - 1] or '').strip() if cols.get('calidad') else ''

            if not marca and not modelo:
                continue

            if any(modelo.upper().startswith(m) for m in ['MÓDULO', 'BATERIA', 'COMPONENTE', 'TAPA', 'PLACA', 'ACCESORIO', 'REPUESTO']) and len(modelo) < 40:
                continue

            def get_num(col_num):
                if not col_num or col_num > len(row):
                    return None
                v = row[col_num - 1]
                if v is None:
                    return None
                try:
                    return float(str(v).replace(',', '.').replace('$', '').strip())
                except:
                    return None

            precio_venta = get_num(cols['venta'])

            # Solo productos con precio de venta válido
            if not precio_venta or precio_venta <= 0:
                continue

            stock = get_num(cols.get('stock')) if cols.get('stock') else None

            key = generar_clave(categoria, marca, modelo, calidad)

            productos.append({
                'source': proveedor,
                'category': categoria,
                'marca': marca,
                'modelo': modelo,
                'calidad': calidad,
                'precio_venta': int(precio_venta),
                'stock': int(stock) if stock else None,
                'key': key
            })

    return productos


def detectar_proveedor(pestana):
    p = pestana.upper()
    if 'ADRICELL' in p:
        return 'ADRICELL'
    elif 'PGVJ' in p:
        return 'PGVJ'
    elif 'PIÑA' in p or 'PINA' in p:
        return 'PINAAPPLE'
    return 'OTROS'


def detectar_categoria(pestana):
    p = pestana.upper()
    if 'MODULO' in p:
        return 'MODULOS'
    elif 'BATERIA' in p:
        return 'BATERIAS'
    elif 'PLACA DE CARGA' in p:
        return 'PLACAS DE CARGA'
    elif 'COMPONENTE' in p:
        return 'COMPONENTES'
    elif 'TAPA' in p:
        return 'TAPAS'
    elif 'PLACA MAIN' in p:
        return 'PLACAS MAIN'
    elif 'REPUESTO' in p:
        return 'REPUESTOS APPLE'
    elif 'ACCESORIO' in p:
        return 'ACCESORIOS'
    return 'OTROS'


def normalizar_texto(s):
    if not s:
        return ''
    s = str(s).lower().strip()
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
    s = re.sub(r'[^a-z0-9]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def generar_clave(categoria, marca, modelo, calidad):
    def n(x):
        return normalizar_texto(x).replace(' ', '_')
    partes = [n(categoria), n(marca), n(modelo), n(calidad)]
    return '_'.join(p for p in partes if p)


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else 'TechnoStore.xlsx'
    salida = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'productos.json')

    print(f"Leyendo Excel: {excel_path}")
    productos = leer_excel(excel_path)
    print(f"Productos generados: {len(productos)}")

    with open(salida, 'w', encoding='utf-8') as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)

    print(f"Guardado: {os.path.abspath(salida)}")
    print(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")


if __name__ == '__main__':
    main()