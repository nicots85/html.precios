"""
SincronizaciÃ³n Excel maestro â†’ Firestore.
Lee TechnoStore.xlsx y actualiza la base de datos con costo y stock por producto.

Uso:
    python backend/sync_excel_firestore.py

Requisitos:
    pip install firebase-admin openpyxl
    Descargar clave de servicio desde Firebase Console
"""
import json
import os
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')


def get_firebase_credentials():
    """Obtiene credenciales de Firebase desde variable de entorno o archivo."""
    creds_path = os.environ.get('FIREBASE_CREDENTIALS_PATH', 'backend/firebase-credentials.json')
    
    if not os.path.exists(creds_path):
        print(f"ERROR: No se encontrÃ³ el archivo de credenciales en: {creds_path}")
        print()
        print("Para obtenerlo:")
        print("1. AndÃ¡ a Firebase Console â†’ Proyecto 'technostore-arg'")
        print("2. ConfiguraciÃ³n del proyecto â†’ Cuentas de servicio")
        print("3. 'Generar nueva clave privada'")
        print("4. GuardÃ¡ el JSON como: backend/firebase-credentials.json")
        sys.exit(1)
    
    return creds_path


def init_firebase():
    """Inicializa Firebase Admin SDK."""
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        print("ERROR: InstalÃ¡ firebase-admin:")
        print("  pip install firebase-admin")
        sys.exit(1)
    
    creds_path = get_firebase_credentials()
    cred = credentials.Certificate(creds_path)
    
    try:
        firebase_admin.get_app()
    except ValueError:
        firebase_admin.initialize_app(cred)
    
    return firestore.client()


def leer_excel(ruta_excel):
    """Lee el Excel maestro y devuelve productos por proveedor y pestaÃ±a."""
    import openpyxl
    
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    productos_por_pestana = {}
    
    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        
        # Detectar fila de encabezado
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val and ('MARCA' in str(val).upper() or 'MODELO' in str(val).upper()):
                    header_row = r
                    break
            if header_row:
                break
        
        if not header_row:
            continue
        
        # Detectar columnas
        cols = {}
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=c).value).strip().upper() if ws.cell(row=header_row, column=c).value else ''
            if 'MARCA' in val:
                cols['marca'] = c
            elif 'MODELO' in val or 'DESCRIPCION' in val:
                cols['modelo'] = c
            elif 'CALIDAD' in val or 'TIPO' in val:
                cols['calidad'] = c
            elif 'VENTA' in val:
                cols['venta'] = c
            elif 'PRECIO PESOS' in val or ('PESOS' in val and 'PRECIO' in val):
                cols['costo_pesos'] = c
            elif 'COSTO' in val and 'ARS' in val:
                cols['costo_pesos'] = c
            elif 'PRECIO DOLAR' in val or 'COSTO' in val and 'U$S' in val:
                cols['costo_usd'] = c
            elif 'DOLAR' in val or 'U$S' in val:
                cols['costo_usd'] = c
            elif 'STOCK' in val:
                cols['stock'] = c
        
        # Leer productos
        productos = []
        for row_num in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
            
            marca = str(row[cols.get('marca', 1) - 1] or '').strip() if cols.get('marca') else ''
            modelo = str(row[cols.get('modelo', 2) - 1] or '').strip() if cols.get('modelo') else ''
            calidad = str(row[cols.get('calidad', 3) - 1] or '').strip() if cols.get('calidad') else ''
            
            if not marca and not modelo:
                continue
            
            # Saltar subtÃ­tulos
            if any(modelo.upper().startswith(m) for m in ['MÃ“DULO', 'BATERIA', 'COMPONENTE', 'TAPA', 'PLACA', 'ACCESORIO', 'REPUESTO']) and len(modelo) < 40:
                continue
            
            def get_num(col_num):
                if not col_num or col_num > len(row):
                    return None
                v = row[col_num - 1]
                if v is None:
                    return None
                try:
                    return float(str(v).replace(',', '.').replace('$', '').strip())
                except:
                    return None
            
            costo_pesos = get_num(cols.get('costo_pesos'))
            costo_usd = get_num(cols.get('costo_usd'))
            precio_venta = get_num(cols.get('venta'))
            stock = get_num(cols.get('stock'))
            
            # Si no hay costo_pesos, intentar de venta
            if costo_pesos is None and precio_venta:
                costo_pesos = precio_venta
            
            productos.append({
                'marca': marca,
                'modelo': modelo,
                'calidad': calidad,
                'costo_pesos': costo_pesos,
                'costo_usd': costo_usd,
                'precio_venta': precio_venta,
                'stock': int(stock) if stock else None
            })
        
        if productos:
            productos_por_pestana[tab_name] = productos
    
    return productos_por_pestana


def detectar_proveedor(pestana):
    """Detecta el proveedor de una pestaÃ±a."""
    p = pestana.upper()
    if 'ADRICELL' in p:
        return 'ADRICELL'
    elif 'PGVJ' in p:
        return 'PGVJ'
    elif 'PIÃ‘A' in p or 'PINA' in p:
        return 'PINAAPPLE'
    return 'OTROS'


def detectar_categoria(pestana):
    """Detecta la categorÃ­a del producto."""
    p = pestana.upper()
    if 'MODULO' in p:
        return 'MODULOS'
    elif 'BATERIA' in p:
        return 'BATERIAS'
    elif 'PLACA DE CARGA' in p or 'PLACAS DE CARGA' in p:
        return 'PLACAS DE CARGA'
    elif 'COMPONENTE' in p or 'PARTES VARIAS' in p:
        return 'COMPONENTES'
    elif 'TAPA' in p:
        return 'TAPAS'
    elif 'PLACA MAIN' in p:
        return 'PLACAS MAIN'
    elif 'REPUESTO' in p:
        return 'REPUESTOS APPLE'
    elif 'ACCESORIO' in p:
        return 'ACCESORIOS'
    elif 'TEMPLADO' in p:
        return 'TEMPLADOS'
    return 'OTROS'


def generar_doc_id(proveedor, marca, modelo, calidad):
    """Genera un ID de documento Ãºnico para Firestore."""
    def normalize(s):
        if not s:
            return ''
        s = str(s).lower().strip()
        import unicodedata
        s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
        import re
        s = re.sub(r'[^a-z0-9]', '_', s)
        s = re.sub(r'_+', '_', s)
        return s.strip('_')
    
    partes = [proveedor, marca, modelo, calidad]
    return '_'.join(normalize(p) for p in partes if p)


def sync_a_firestore(excel_path='TechnoStore.xlsx', batch_size=500):
    """Sincroniza el Excel con Firestore."""
    
    print(f"Leyendo Excel: {excel_path}")
    productos_por_pestana = leer_excel(excel_path)
    
    total_productos = sum(len(p) for p in productos_por_pestana.values())
    print(f"Productos encontrados: {total_productos}")
    print()
    
    print("Inicializando Firestore...")
    db = init_firebase()
    
    timestamp = datetime.utcnow().isoformat()
    batch = db.batch()
    contador = 0
    
    for pestana, productos in productos_por_pestana.items():
        proveedor = detectar_proveedor(pestana)
        categoria = detectar_categoria(pestana)
        
        print(f"\n[{proveedor}] {pestana} ({len(productos)} productos)")
        
        for prod in productos:
            doc_id = generar_doc_id(proveedor, prod['marca'], prod['modelo'], prod['calidad'])
            
            doc_data = {
                'proveedor': proveedor,
                'categoria': categoria,
                'pestana': pestana,
                'marca': prod['marca'],
                'modelo': prod['modelo'],
                'calidad': prod['calidad'],
                'costo_pesos': prod['costo_pesos'],
                'costo_usd': prod['costo_usd'],
                'precio_venta': prod['precio_venta'],
                'stock': prod['stock'],
                'updated_at': timestamp
            }
            
            doc_ref = db.collection('productos').document(doc_id)
            batch.set(doc_ref, doc_data, merge=True)
            contador += 1
            
            if contador >= batch_size:
                batch.commit()
                print(f"  Commit batch: {contador} productos")
                batch = db.batch()
                contador = 0
    
    if contador > 0:
        batch.commit()
        print(f"  Commit final: {contador} productos")
    
    print()
    print(f"SincronizaciÃ³n completa: {total_productos} productos en Firestore")
    return total_productos


if __name__ == '__main__':
    excel_path = sys.argv[1] if len(sys.argv) > 1 else 'TechnoStore.xlsx'
    sync_a_firestore(excel_path)
